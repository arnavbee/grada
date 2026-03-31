import csv
import io
import json
import re
import zlib
from difflib import SequenceMatcher
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from PIL import Image

from app.services.object_storage import get_object_storage_service

DOCUMENT_TYPE_CATALOG = 'catalog'
DOCUMENT_TYPE_PO_BUILDER = 'po_builder'
DOCUMENT_TYPE_PACKING_LIST = 'packing_list'
DOCUMENT_TYPE_BARCODE = 'barcode'
TEMPLATE_KIND_TABULAR = 'tabular'
TEMPLATE_KIND_WORKBOOK = 'workbook'
TEMPLATE_KIND_PDF_LAYOUT = 'pdf_layout'
TEMPLATE_KIND_STICKER = 'sticker'

MARKETPLACE_LABELS = {
    'generic': 'Generic',
    'myntra': 'Myntra',
    'ajio': 'Ajio',
    'amazon_in': 'Amazon IN',
    'flipkart': 'Flipkart',
    'nykaa': 'Nykaa',
    'styli': 'Styli',
    'landmark': 'Landmark',
}

MARKETPLACE_ALIASES = {
    'myntra': 'myntra',
    'ajio': 'ajio',
    'amazon': 'amazon_in',
    'amazon in': 'amazon_in',
    'amazon india': 'amazon_in',
    'amazon_in': 'amazon_in',
    'flipkart': 'flipkart',
    'nykaa': 'nykaa',
    'styli': 'styli',
    'landmark': 'landmark',
    'generic': 'generic',
}

CATALOG_SOURCE_FIELDS = (
    'serial_no',
    'sku',
    'title',
    'brand',
    'category',
    'color',
    'size',
    'mrp',
    'description',
    'fabric',
    'composition',
    'woven_knits',
    'units',
    'po_price',
    'osp',
    'status',
    'image_preview',
    'image_url',
)

PO_BUILDER_SOURCE_FIELDS = (
    'sku_id',
    'brand_name',
    'category_type',
    'styli_sku_id',
    'color',
    'size',
    'l1',
    'fibre_composition',
    'coo',
    'po_price',
    'osp_in_sar',
    'po_qty',
    'knitted_woven',
    'product_name',
    'dress_print',
    'dress_length',
    'dress_shape',
    'sleeve_length',
    'neck_women',
    'sleeve_styling',
)

HEADER_ALIASES = {
    DOCUMENT_TYPE_CATALOG: {
        's no': 'serial_no',
        'serial no': 'serial_no',
        'serial number': 'serial_no',
        'style no': 'sku',
        'style number': 'sku',
        'style id': 'sku',
        'seller sku': 'sku',
        'sellersku': 'sku',
        'sku': 'sku',
        'seller-sku': 'sku',
        'item name': 'title',
        'product name': 'title',
        'name': 'title',
        'title': 'title',
        'brand': 'brand',
        'brand name': 'brand',
        'brand-name': 'brand',
        'category': 'category',
        'department': 'category',
        'item type': 'category',
        'item-type': 'category',
        'color': 'color',
        'color name': 'color',
        'color-name': 'color',
        'shade': 'color',
        'size': 'size',
        'size name': 'size',
        'size-name': 'size',
        'mrp': 'mrp',
        'standard price': 'mrp',
        'standard-price': 'mrp',
        'selling price': 'mrp',
        'description': 'description',
        'product description': 'description',
        'product-description': 'description',
        'fabric': 'fabric',
        'composition': 'composition',
        'woven knits': 'woven_knits',
        'wovenknits': 'woven_knits',
        'woven/knits': 'woven_knits',
        'units': 'units',
        'po price': 'po_price',
        'osp': 'osp',
        'status': 'status',
        'image preview': 'image_preview',
        'image-preview': 'image_preview',
        'primary image url': 'image_url',
        'main image url': 'image_url',
    },
    DOCUMENT_TYPE_PO_BUILDER: {
        'style code': 'sku_id',
        'sku id': 'sku_id',
        'sku': 'sku_id',
        'brand': 'brand_name',
        'brand name': 'brand_name',
        'category': 'category_type',
        'category type': 'category_type',
        'styli sku id': 'styli_sku_id',
        'styli sku': 'styli_sku_id',
        'color': 'color',
        'size': 'size',
        'l1': 'l1',
        'fibre composition': 'fibre_composition',
        'fiber composition': 'fibre_composition',
        'coo': 'coo',
        'po price': 'po_price',
        'osp inside sar': 'osp_in_sar',
        'osp in sar': 'osp_in_sar',
        'po qty': 'po_qty',
        'qty': 'po_qty',
        'knitted woven': 'knitted_woven',
        'product name': 'product_name',
        'dress print': 'dress_print',
        'dress length': 'dress_length',
        'dress shape': 'dress_shape',
        'sleeve length': 'sleeve_length',
        'neck': 'neck_women',
        'neck women': 'neck_women',
        'sleeve styling': 'sleeve_styling',
    },
}

SAMPLE_UPLOAD_DIR = Path('static/uploads/marketplace-templates')
SAMPLE_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
object_storage = get_object_storage_service()


def marketplace_label(marketplace_key: str) -> str:
    return MARKETPLACE_LABELS.get(marketplace_key, marketplace_key.replace('_', ' ').title())


def normalize_marketplace_key(raw_marketplace: str) -> str:
    normalized = re.sub(r'\s+', ' ', raw_marketplace.strip().lower())
    marketplace_key = MARKETPLACE_ALIASES.get(normalized)
    if marketplace_key is not None:
        return marketplace_key
    custom_key = re.sub(r'[^a-z0-9]+', '_', normalized).strip('_')
    if not custom_key:
        allowed = ', '.join(MARKETPLACE_LABELS.values())
        raise ValueError(f'Unsupported marketplace "{raw_marketplace}". Allowed: {allowed}.')
    return custom_key


def source_fields_for_document_type(document_type: str) -> tuple[str, ...]:
    if document_type == DOCUMENT_TYPE_PO_BUILDER:
        return PO_BUILDER_SOURCE_FIELDS
    return CATALOG_SOURCE_FIELDS


def template_kind_for_document_type(document_type: str) -> str:
    if document_type == DOCUMENT_TYPE_PO_BUILDER:
        return TEMPLATE_KIND_WORKBOOK
    if document_type == DOCUMENT_TYPE_PACKING_LIST:
        return TEMPLATE_KIND_PDF_LAYOUT
    if document_type == DOCUMENT_TYPE_BARCODE:
        return TEMPLATE_KIND_STICKER
    return TEMPLATE_KIND_TABULAR


def normalize_template_columns(columns: list[dict[str, object]] | None) -> list[dict[str, object]]:
    normalized: list[dict[str, object]] = []
    seen_headers: set[str] = set()
    for raw_column in columns or []:
        header = str(raw_column.get('header') or '').strip()
        if not header:
            continue
        header_key = header.lower()
        if header_key in seen_headers:
            continue
        seen_headers.add(header_key)
        source_field = str(raw_column.get('source_field') or '').strip() or None
        column: dict[str, object] = {
            'header': header,
            'source_field': source_field,
            'required': bool(raw_column.get('required') or False),
        }
        confidence = raw_column.get('confidence')
        if isinstance(confidence, (int, float)):
            column['confidence'] = max(0.0, min(1.0, float(confidence)))
        normalized.append(column)
    return normalized


def save_sample_file(*, company_id: str, filename: str, content: bytes, content_type: str) -> str:
    extension = Path(filename or '').suffix.lower()
    if extension not in {'.csv', '.xlsx', '.pdf', '.png', '.jpg', '.jpeg', '.webp'}:
        extension = '.bin'
    unique_name = f'{uuid4().hex}{extension}'
    key = f'uploads/{company_id}/marketplace-templates/{unique_name}'
    if object_storage.enabled:
        object_url = object_storage.upload_bytes(key=key, content=content, content_type=content_type)
        if object_url:
            return object_url
    company_dir = SAMPLE_UPLOAD_DIR / company_id
    company_dir.mkdir(parents=True, exist_ok=True)
    output_path = company_dir / unique_name
    output_path.write_bytes(content)
    return f'/static/uploads/marketplace-templates/{company_id}/{unique_name}'


def parse_marketplace_sample(*, content: bytes, filename: str, document_type: str) -> dict[str, object]:
    file_format = _detect_file_format(filename)
    if document_type == DOCUMENT_TYPE_PACKING_LIST:
        return _parse_packing_list_sample(content, file_format=file_format)
    if document_type == DOCUMENT_TYPE_BARCODE:
        return _parse_barcode_sample(content, file_format=file_format)
    if file_format == 'xlsx':
        return _parse_xlsx_sample(content, document_type=document_type)
    return _parse_csv_sample(content, document_type=document_type)


def _detect_file_format(filename: str) -> str:
    extension = Path(filename or '').suffix.lower()
    if extension == '.xlsx':
        return 'xlsx'
    if extension == '.pdf':
        return 'pdf'
    if extension in {'.png', '.jpg', '.jpeg', '.webp'}:
        return extension.lstrip('.')
    return 'csv'


def _parse_packing_list_sample(content: bytes, *, file_format: str) -> dict[str, object]:
    if file_format != 'pdf':
        raise ValueError('Packing list learning currently supports PDF samples only.')
    text_candidates = _extract_pdf_text_candidates(content)
    title = next(
        (
            candidate
            for candidate in text_candidates
            if 'packing' in candidate.lower() and 'list' in candidate.lower()
        ),
        'PACKING LIST',
    )
    po_label = next(
        (
            candidate
            for candidate in text_candidates
            if re.search(r'\bpo\b.*\b(no|number)\b', candidate, flags=re.IGNORECASE)
        ),
        'PO NUMBER',
    )
    qty_label = next(
        (
            candidate
            for candidate in text_candidates
            if re.search(r'\b(qty|quantity)\b', candidate, flags=re.IGNORECASE)
        ),
        'Qty',
    )
    detected_headers = [title, po_label, qty_label]
    return {
        'file_format': file_format,
        'sheet_name': None,
        'header_row_index': 1,
        'detected_headers': detected_headers,
        'columns': [],
        'layout': {
            'layout_key': 'default_v1',
            'title': title,
            'meta_labels': {
                'po_number': po_label,
            },
            'column_headers': {
                'quantity': qty_label,
            },
            'learned_from_sample': True,
        },
    }


def _parse_barcode_sample(content: bytes, *, file_format: str) -> dict[str, object]:
    if file_format == 'pdf':
        width_mm, height_mm = _extract_pdf_dimensions_mm(content)
        return {
            'file_format': file_format,
            'sheet_name': None,
            'header_row_index': 1,
            'detected_headers': [],
            'columns': [],
            'layout': {
                'sticker_template_kind': 'styli',
                'sticker_template_id': None,
                'width_mm': width_mm,
                'height_mm': height_mm,
                'measurement_source': 'pdf_page_box',
                'learned_from_sample': True,
            },
        }
    if file_format in {'png', 'jpg', 'jpeg', 'webp'}:
        width_mm, height_mm, source = _extract_image_dimensions_mm(content)
        return {
            'file_format': file_format,
            'sheet_name': None,
            'header_row_index': 1,
            'detected_headers': [],
            'columns': [],
            'layout': {
                'sticker_template_kind': 'styli',
                'sticker_template_id': None,
                'width_mm': width_mm,
                'height_mm': height_mm,
                'measurement_source': source,
                'learned_from_sample': True,
            },
        }
    raise ValueError('Barcode learning currently supports PDF and image samples only.')


def _parse_csv_sample(content: bytes, *, document_type: str) -> dict[str, object]:
    decoded = content.decode('utf-8-sig', errors='ignore')
    rows = list(csv.reader(io.StringIO(decoded)))
    header_row_index = _detect_header_row_index(rows, document_type=document_type)
    detected_headers = _extract_headers_from_row(rows[header_row_index - 1] if rows else [])
    return {
        'file_format': 'csv',
        'sheet_name': None,
        'header_row_index': header_row_index,
        'detected_headers': detected_headers,
        'columns': _build_mapping_suggestions(detected_headers, document_type=document_type),
        'layout': {'prefix_rows': rows[:header_row_index]},
    }


def _parse_xlsx_sample(content: bytes, *, document_type: str) -> dict[str, object]:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    rows: list[list[str]] = []
    max_preview_rows = min(sheet.max_row, 10)
    max_preview_cols = min(sheet.max_column, 80)
    for row in sheet.iter_rows(min_row=1, max_row=max_preview_rows, max_col=max_preview_cols, values_only=True):
        rows.append([_cell_to_string(value) for value in row])
    header_row_index = _detect_header_row_index(rows, document_type=document_type)
    detected_headers = _extract_headers_from_row(rows[header_row_index - 1] if rows else [])
    layout = {
        'freeze_panes': sheet.freeze_panes if isinstance(sheet.freeze_panes, str) else None,
        'merged_ranges': [str(cell_range) for cell_range in sheet.merged_cells.ranges][:50],
        'column_widths': {
            column_letter: dimension.width
            for column_letter, dimension in sheet.column_dimensions.items()
            if dimension.width is not None
        },
        'prefix_rows': rows[:header_row_index],
    }
    return {
        'file_format': 'xlsx',
        'sheet_name': sheet.title,
        'header_row_index': header_row_index,
        'detected_headers': detected_headers,
        'columns': _build_mapping_suggestions(detected_headers, document_type=document_type),
        'layout': layout,
    }


def _cell_to_string(value: object) -> str:
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _extract_pdf_text_candidates(content: bytes) -> list[str]:
    chunks = [content, *_extract_pdf_stream_chunks(content)]
    candidates: list[str] = []
    seen: set[str] = set()
    for chunk in chunks:
        for match in re.finditer(rb'\(([^()]*)\)', chunk):
            cleaned = _clean_pdf_text(match.group(1))
            if cleaned and cleaned.lower() not in seen:
                seen.add(cleaned.lower())
                candidates.append(cleaned)
    return candidates


def _extract_pdf_stream_chunks(content: bytes) -> list[bytes]:
    chunks: list[bytes] = []
    for match in re.finditer(rb'stream\r?\n(.*?)\r?\nendstream', content, flags=re.DOTALL):
        stream_bytes = match.group(1)
        if not stream_bytes:
            continue
        chunks.append(stream_bytes)
        try:
            chunks.append(zlib.decompress(stream_bytes))
        except Exception:
            continue
    return chunks


def _clean_pdf_text(raw: bytes) -> str:
    text = raw.decode('latin-1', errors='ignore')
    text = text.replace(r'\(', '(').replace(r'\)', ')').replace(r'\\', '\\')
    text = re.sub(r'\\[0-7]{1,3}', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    if len(text) < 2 or not re.search(r'[A-Za-z]', text):
        return ''
    return text


def _extract_pdf_dimensions_mm(content: bytes) -> tuple[float, float]:
    match = re.search(
        rb'/MediaBox\s*\[\s*([0-9.\-]+)\s+([0-9.\-]+)\s+([0-9.\-]+)\s+([0-9.\-]+)\s*\]',
        content,
    )
    if match is None:
        return 45.03, 60.0
    x1, y1, x2, y2 = [float(value) for value in match.groups()]
    width_pt = max(1.0, x2 - x1)
    height_pt = max(1.0, y2 - y1)
    return round(width_pt * 25.4 / 72.0, 2), round(height_pt * 25.4 / 72.0, 2)


def _extract_image_dimensions_mm(content: bytes) -> tuple[float, float, str]:
    image = Image.open(io.BytesIO(content))
    width_px, height_px = image.size
    dpi_info = image.info.get('dpi')
    if (
        isinstance(dpi_info, tuple)
        and len(dpi_info) >= 2
        and dpi_info[0]
        and dpi_info[1]
        and dpi_info[0] > 0
        and dpi_info[1] > 0
    ):
        x_dpi = float(dpi_info[0])
        y_dpi = float(dpi_info[1])
        source = 'image_dpi'
    else:
        x_dpi = 203.0
        y_dpi = 203.0
        source = 'image_default_203dpi'
    width_mm = round(width_px * 25.4 / x_dpi, 2)
    height_mm = round(height_px * 25.4 / y_dpi, 2)
    return width_mm, height_mm, source


def _detect_header_row_index(rows: list[list[str]], *, document_type: str) -> int:
    best_index = 1
    best_score = -1.0
    for index, row in enumerate(rows[:10], start=1):
        score = _score_header_row(row, document_type=document_type)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _score_header_row(row: list[str], *, document_type: str) -> float:
    headers = _extract_headers_from_row(row)
    if not headers:
        return -1.0
    alias_hits = 0
    alpha_hits = 0
    for header in headers:
        suggestion = _suggest_source_field(header, document_type=document_type)
        if suggestion['source_field']:
            alias_hits += 1
        if re.search(r'[A-Za-z]', header):
            alpha_hits += 1
    unique_headers = len({_normalize_header_token(header) for header in headers})
    return alias_hits * 100 + alpha_hits * 10 + unique_headers


def _extract_headers_from_row(row: list[str]) -> list[str]:
    headers: list[str] = []
    seen_non_empty = False
    for value in row:
        cleaned = value.strip()
        if cleaned:
            seen_non_empty = True
            headers.append(cleaned)
            continue
        if seen_non_empty:
            break
    return headers


def _build_mapping_suggestions(headers: list[str], *, document_type: str) -> list[dict[str, object]]:
    return [_suggest_source_field(header, document_type=document_type) for header in headers]


def _suggest_source_field(header: str, *, document_type: str) -> dict[str, object]:
    normalized_header = _normalize_header_token(header)
    aliases = HEADER_ALIASES.get(document_type, {})
    direct_match = aliases.get(normalized_header)
    if direct_match:
        return {'header': header, 'source_field': direct_match, 'required': False, 'confidence': 0.98}

    compact_header = normalized_header.replace(' ', '')
    for alias_key, source_field in aliases.items():
        if compact_header == alias_key.replace(' ', ''):
            return {'header': header, 'source_field': source_field, 'required': False, 'confidence': 0.95}
        if compact_header and compact_header in alias_key.replace(' ', ''):
            return {'header': header, 'source_field': source_field, 'required': False, 'confidence': 0.72}

    best_field = None
    best_score = 0.0
    for source_field in source_fields_for_document_type(document_type):
        comparison = source_field.replace('_', ' ')
        score = SequenceMatcher(None, normalized_header, comparison).ratio()
        if score > best_score:
            best_score = score
            best_field = source_field
    if best_field and best_score >= 0.6:
        return {'header': header, 'source_field': best_field, 'required': False, 'confidence': round(best_score, 2)}
    return {'header': header, 'source_field': None, 'required': False, 'confidence': 0.0}


def _normalize_header_token(value: str) -> str:
    cleaned = re.sub(r'[^a-z0-9]+', ' ', value.strip().lower())
    return re.sub(r'\s+', ' ', cleaned).strip()


def dumps_json(payload: dict[str, object]) -> str:
    return json.dumps(payload, separators=(',', ':'))


def loads_json(raw: str | None) -> dict[str, object]:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}
