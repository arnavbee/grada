import json
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config.styli_attributes import (
    DEFAULT_SIZE_RATIO,
    DRESS_ATTRIBUTES,
    LOW_CONFIDENCE_THRESHOLD,
    PO_EXPORT_GROUP_ROW,
    PO_EXPORT_HEADERS,
    SIZE_ORDER,
)
from app.models import Company, PORequest, PORequestColorway, PORequestItem, PORequestRow, Product, ProductImage


DEFAULT_PO_ITEM_ATTRIBUTES = {
    'fields': {},
    'review_required': False,
}


def _json_loads(raw: str | None) -> dict[str, Any]:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def normalize_size_ratio(raw_ratio: dict[str, Any] | None) -> dict[str, int]:
    normalized = dict(DEFAULT_SIZE_RATIO)
    if not isinstance(raw_ratio, dict):
        return normalized

    for size in SIZE_ORDER:
        raw_value = raw_ratio.get(size)
        try:
            normalized[size] = max(0, int(raw_value))
        except (TypeError, ValueError):
            continue
    return normalized


def _normalize_letter(raw_letter: str | None, fallback_index: int) -> str:
    if raw_letter:
        token = ''.join(ch for ch in raw_letter.strip().upper() if ch.isalpha())
        if token:
            return token[:2]
    first = fallback_index % 26
    second = fallback_index // 26
    if second <= 0:
        return chr(65 + first)
    return f'{chr(64 + second)}{chr(65 + first)}'


def normalize_color_name(raw_color_name: str | None, default_color: str | None = None) -> str:
    token = (raw_color_name or default_color or 'Default').strip()
    return token or 'Default'


def _sanitize_color_token(raw_value: str) -> str:
    cleaned = ''.join(ch if ch.isalnum() else '_' for ch in raw_value.strip().upper())
    collapsed = '_'.join(part for part in cleaned.split('_') if part)
    return collapsed or 'DEFAULT'


def generate_row_sku(base_style_code: str | None, colorway_letter: str, color_name: str, size: str) -> str:
    base = (base_style_code or '').strip().upper()
    if not base:
        return ''
    return f'{base}-{colorway_letter}-{_sanitize_color_token(color_name)}-{size.upper()}'


def _coerce_decimal(raw_value: Any) -> float | None:
    if raw_value in (None, ''):
        return None
    try:
        return float(Decimal(str(raw_value)))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _normalize_ai_fields(raw_attributes: dict[str, Any] | None) -> dict[str, Any]:
    if not isinstance(raw_attributes, dict):
        return dict(DEFAULT_PO_ITEM_ATTRIBUTES)

    raw_fields = raw_attributes.get('fields')
    if isinstance(raw_fields, dict):
        fields: dict[str, dict[str, Any]] = {}
        review_required = bool(raw_attributes.get('review_required'))
        for field_name, raw_field in raw_fields.items():
            if isinstance(raw_field, dict):
                value = str(raw_field.get('value') or '').strip()
                confidence = _coerce_decimal(raw_field.get('confidence'))
            else:
                value = str(raw_field or '').strip()
                confidence = None
            fields[field_name] = {
                'value': value,
                'confidence': confidence,
            }
            if confidence is not None and confidence < LOW_CONFIDENCE_THRESHOLD:
                review_required = True
        return {
            'fields': fields,
            'review_required': review_required,
        }

    fields = {}
    review_required = False
    for field_name, raw_value in raw_attributes.items():
        if isinstance(raw_value, dict):
            value = str(raw_value.get('value') or '').strip()
            confidence = _coerce_decimal(raw_value.get('confidence'))
        else:
            value = str(raw_value or '').strip()
            confidence = None
        fields[field_name] = {
            'value': value,
            'confidence': confidence,
        }
        if confidence is not None and confidence < LOW_CONFIDENCE_THRESHOLD:
            review_required = True
    return {
        'fields': fields,
        'review_required': review_required,
    }


def normalize_ai_attributes(raw_attributes: dict[str, Any] | None) -> dict[str, Any]:
    normalized = _normalize_ai_fields(raw_attributes)
    fields = normalized['fields']

    # Keep dress attributes constrained to known Styli enums whenever possible.
    for field_name, allowed_values in DRESS_ATTRIBUTES.items():
        raw_field = fields.get(field_name) or {'value': '', 'confidence': None}
        value = str(raw_field.get('value') or '').strip()
        if value and value not in allowed_values:
            matched = next((candidate for candidate in allowed_values if candidate.lower() == value.lower()), '')
            if matched:
                value = matched
        fields[field_name] = {
            'value': value,
            'confidence': _coerce_decimal(raw_field.get('confidence')),
        }

    if 'woven_knits' not in fields:
        fields['woven_knits'] = {'value': 'Woven', 'confidence': None}

    normalized['review_required'] = normalized['review_required'] or any(
        (
            isinstance(field_payload, dict)
            and field_payload.get('confidence') is not None
            and float(field_payload['confidence']) < LOW_CONFIDENCE_THRESHOLD
        )
        for field_payload in fields.values()
    )
    return normalized


def get_attribute_value(raw_attributes: dict[str, Any] | None, field_name: str, fallback: str = '') -> str:
    normalized = normalize_ai_attributes(raw_attributes)
    field_payload = normalized['fields'].get(field_name)
    if not isinstance(field_payload, dict):
        return fallback
    value = str(field_payload.get('value') or '').strip()
    return value or fallback


def ensure_item_colorways(item: PORequestItem, default_color: str | None = None) -> list[PORequestColorway]:
    existing = sorted(item.colorways, key=lambda colorway: (colorway.letter, colorway.created_at))
    if existing:
        return existing

    colorway = PORequestColorway(
        letter='A',
        color_name=normalize_color_name(default_color),
    )
    item.colorways.append(colorway)
    return [colorway]


def apply_item_updates(
    item: PORequestItem,
    payload: Any,
    *,
    default_color: str | None = None,
) -> None:
    if payload.po_price is not None:
        item.po_price = payload.po_price
    if payload.osp_inside_price is not None:
        item.osp_inside_price = payload.osp_inside_price
    if payload.fabric_composition is not None:
        item.fabric_composition = payload.fabric_composition.strip()
    if payload.size_ratio is not None:
        item.size_ratio = normalize_size_ratio(payload.size_ratio)
    if payload.extracted_attributes is not None:
        item.extracted_attributes = normalize_ai_attributes(payload.extracted_attributes)

    if payload.colorways is not None:
        item.colorways.clear()
        for index, colorway_payload in enumerate(payload.colorways):
            item.colorways.append(
                PORequestColorway(
                    letter=_normalize_letter(colorway_payload.letter, index),
                    color_name=normalize_color_name(colorway_payload.color_name, default_color),
                )
            )

    ensure_item_colorways(item, default_color)
    item.size_ratio = normalize_size_ratio(item.size_ratio)
    item.extracted_attributes = normalize_ai_attributes(item.extracted_attributes)


def _derive_fabric_short_name(fabric_composition: str | None, product: Product) -> str:
    raw_value = (fabric_composition or '').strip()
    if not raw_value:
        ai_attributes = _json_loads(product.ai_attributes_json)
        raw_value = str(ai_attributes.get('fabric') or product.title or '').strip()
    if not raw_value:
        return 'Fabric'

    cleaned = raw_value.split(',')[0].strip()
    lowered = cleaned.lower()
    replacements = {
        '100% polyester': 'Polyester',
        'polyester': 'Polyester',
        'cotton': 'Cotton',
        'viscose': 'Viscose',
        'rayon': 'Rayon',
        'linen': 'Linen',
    }
    for token, replacement in replacements.items():
        if token in lowered:
            return replacement
    words = [word for word in cleaned.replace('%', '').split() if word]
    return words[-1].title() if words else 'Fabric'


def generate_product_name(
    product: Product,
    *,
    color_name: str,
    fabric_composition: str | None,
    attributes: dict[str, Any] | None,
) -> str:
    category_name = (product.category or 'Dresses').strip()
    lowered_category = category_name.lower()
    singular_map = {
        'dresses': 'Dress',
        'tops': 'Top',
        'shirts': 'Shirt',
        'trousers': 'Trouser',
        'sets': 'Set',
    }
    category_token = singular_map.get(lowered_category, category_name[:-1] if category_name.endswith('s') else category_name)
    if not category_token:
        category_token = 'Dress'

    length = get_attribute_value(attributes, 'dress_length')
    title_tokens = ['Women\'s', color_name.strip() or product.color or 'Default']
    fabric_short_name = _derive_fabric_short_name(fabric_composition, product)
    if fabric_short_name:
        title_tokens.append(fabric_short_name)
    if length:
        title_tokens.append(length)
    title_tokens.append(category_token.title())
    title_tokens.append(category_name.upper())
    return ' '.join(token for token in title_tokens if token)


@dataclass
class SerializedProductSnapshot:
    id: str
    sku: str
    title: str
    brand: str
    category: str
    color: str
    primary_image_url: str | None


def _load_product_snapshots(db: Session, po_request: PORequest) -> dict[str, SerializedProductSnapshot]:
    product_ids = [item.product_id for item in po_request.items]
    if not product_ids:
        return {}

    products = db.scalars(select(Product).where(Product.id.in_(product_ids))).all()
    image_rows = db.scalars(
        select(ProductImage).where(ProductImage.product_id.in_(product_ids)).order_by(ProductImage.created_at.asc())
    ).all()
    image_map: dict[str, str] = {}
    for image_row in image_rows:
        image_map.setdefault(image_row.product_id, image_row.file_url)

    snapshots: dict[str, SerializedProductSnapshot] = {}
    for product in products:
        ai_payload = _json_loads(product.ai_attributes_json)
        snapshots[product.id] = SerializedProductSnapshot(
            id=product.id,
            sku=product.sku,
            title=product.title,
            brand=product.brand or '',
            category=product.category or '',
            color=product.color or '',
            primary_image_url=image_map.get(product.id) or (
                str(ai_payload.get('primary_image_url')) if ai_payload.get('primary_image_url') else None
            ),
        )
    return snapshots


def rebuild_po_request_rows(db: Session, po_request: PORequest) -> list[PORequestRow]:
    db.flush()
    db.query(PORequestRow).filter(PORequestRow.po_request_id == po_request.id).delete()

    company = db.query(Company).filter(Company.id == po_request.company_id).first()
    company_name = company.name if company else 'Brand'
    products_by_id = {
        product.id: product
        for product in db.scalars(
            select(Product).where(Product.id.in_([item.product_id for item in po_request.items]))
        ).all()
    }

    rows: list[PORequestRow] = []
    row_index = 1
    for item in po_request.items:
        product = products_by_id.get(item.product_id)
        if product is None:
            continue

        ensure_item_colorways(item, product.color)
        item.size_ratio = normalize_size_ratio(item.size_ratio)
        item.extracted_attributes = normalize_ai_attributes(item.extracted_attributes)
        colorways = sorted(item.colorways, key=lambda colorway: (colorway.letter, colorway.created_at))

        for colorway in colorways:
            color_name = normalize_color_name(colorway.color_name, product.color)
            for size in SIZE_ORDER:
                quantity = item.size_ratio.get(size, 0)
                if quantity <= 0:
                    continue

                row = PORequestRow(
                    po_request_id=po_request.id,
                    po_request_item_id=item.id,
                    product_id=product.id,
                    row_index=row_index,
                    sku_id=generate_row_sku(product.sku, colorway.letter, color_name, size),
                    brand_name=(product.brand or company_name).strip() or company_name,
                    category_type=(product.category or 'Dresses').strip() or 'Dresses',
                    styli_sku_id='',
                    color=color_name,
                    size=size,
                    colorway_letter=colorway.letter,
                    fibre_composition=(item.fabric_composition or '').strip(),
                    coo='India',
                    po_price=item.po_price,
                    osp_in_sar=item.osp_inside_price,
                    po_qty=quantity,
                    knitted_woven=get_attribute_value(item.extracted_attributes, 'woven_knits', 'Woven') or 'Woven',
                    product_name=generate_product_name(
                        product,
                        color_name=color_name,
                        fabric_composition=item.fabric_composition,
                        attributes=item.extracted_attributes,
                    ),
                    dress_print=get_attribute_value(item.extracted_attributes, 'dress_print'),
                    dress_length=get_attribute_value(item.extracted_attributes, 'dress_length'),
                    dress_shape=get_attribute_value(item.extracted_attributes, 'dress_shape'),
                    sleeve_length=get_attribute_value(item.extracted_attributes, 'sleeve_length'),
                    neck_women=get_attribute_value(item.extracted_attributes, 'neck_women'),
                    sleeve_styling=get_attribute_value(item.extracted_attributes, 'sleeve_styling'),
                    other_attributes_json='{}',
                )
                rows.append(row)
                db.add(row)
                row_index += 1

    db.flush()
    return rows


def serialize_po_request(db: Session, po_request: PORequest) -> dict[str, Any]:
    product_snapshots = _load_product_snapshots(db, po_request)
    items = sorted(po_request.items, key=lambda item: item.created_at)
    rows = sorted(po_request.rows, key=lambda row: (row.row_index, row.created_at))

    item_payloads: list[dict[str, Any]] = []
    for item in items:
        snapshot = product_snapshots.get(item.product_id)
        item_payloads.append(
            {
                'id': item.id,
                'po_request_id': item.po_request_id,
                'product_id': item.product_id,
                'po_price': float(item.po_price) if item.po_price is not None else None,
                'osp_inside_price': float(item.osp_inside_price) if item.osp_inside_price is not None else None,
                'fabric_composition': item.fabric_composition,
                'size_ratio': normalize_size_ratio(item.size_ratio),
                'extracted_attributes': normalize_ai_attributes(item.extracted_attributes),
                'colorways': [
                    {
                        'id': colorway.id,
                        'po_request_item_id': colorway.po_request_item_id,
                        'letter': colorway.letter,
                        'color_name': colorway.color_name,
                        'created_at': colorway.created_at,
                        'updated_at': colorway.updated_at,
                    }
                    for colorway in sorted(item.colorways, key=lambda colorway: (colorway.letter, colorway.created_at))
                ],
                'product': (
                    {
                        'id': snapshot.id,
                        'sku': snapshot.sku,
                        'title': snapshot.title,
                        'brand': snapshot.brand,
                        'category': snapshot.category,
                        'color': snapshot.color,
                        'primary_image_url': snapshot.primary_image_url,
                    }
                    if snapshot
                    else None
                ),
                'created_at': item.created_at,
                'updated_at': item.updated_at,
            }
        )

    row_payloads = [
        {
            'id': row.id,
            'po_request_id': row.po_request_id,
            'po_request_item_id': row.po_request_item_id,
            'product_id': row.product_id,
            'row_index': row.row_index,
            'sku_id': row.sku_id,
            'brand_name': row.brand_name,
            'category_type': row.category_type,
            'styli_sku_id': row.styli_sku_id,
            'color': row.color,
            'size': row.size,
            'colorway_letter': row.colorway_letter,
            'l1': row.l1,
            'fibre_composition': row.fibre_composition,
            'coo': row.coo,
            'po_price': float(row.po_price) if row.po_price is not None else None,
            'osp_in_sar': float(row.osp_in_sar) if row.osp_in_sar is not None else None,
            'po_qty': row.po_qty,
            'knitted_woven': row.knitted_woven,
            'product_name': row.product_name,
            'dress_print': row.dress_print,
            'dress_length': row.dress_length,
            'dress_shape': row.dress_shape,
            'sleeve_length': row.sleeve_length,
            'neck_women': row.neck_women,
            'sleeve_styling': row.sleeve_styling,
            'created_at': row.created_at,
            'updated_at': row.updated_at,
        }
        for row in rows
    ]

    return {
        'id': po_request.id,
        'company_id': po_request.company_id,
        'status': po_request.status,
        'created_by_user_id': po_request.created_by_user_id,
        'created_at': po_request.created_at,
        'updated_at': po_request.updated_at,
        'items': item_payloads,
        'rows': row_payloads,
    }


def _format_cell_reference(column_index: int, row_index: int) -> str:
    column_name = ''
    working_index = column_index
    while working_index > 0:
        working_index, remainder = divmod(working_index - 1, 26)
        column_name = chr(65 + remainder) + column_name
    return f'{column_name}{row_index}'

HEADER_FILL_COLORS = {
    'Core': 'D9EAD3',
    'Tops': 'D0E0E3',
    'Trousers': 'FFF2CC',
    'Dress': 'EAD1DC',
    'Coord Set / Jumpsuits': 'FCE5CD',
    'Ethnic Kurti / Set': '9FD5F2',
    'Denim Jeans': 'EADCF8',
    'Outerwear / Jackets': 'F4E1D2',
}

THIN_SIDE = Side(style='thin', color='000000')
HEADER_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_BODY_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _po_export_field_map(row: PORequestRow) -> dict[str, Any]:
    return {
        'sku_id': row.sku_id,
        'brand_name': row.brand_name,
        'category_type': row.category_type,
        'styli_sku_id': row.styli_sku_id or '',
        'color': row.color,
        'size': row.size,
        'l1': row.l1,
        'fibre_composition': row.fibre_composition or '',
        'coo': row.coo,
        'po_price': float(row.po_price) if row.po_price is not None else '',
        'osp_in_sar': float(row.osp_in_sar) if row.osp_in_sar is not None else '',
        'po_qty': row.po_qty,
        'knitted_woven': row.knitted_woven or '',
        'product_name': row.product_name or '',
        'dress_print': row.dress_print or '',
        'dress_length': row.dress_length or '',
        'dress_shape': row.dress_shape or '',
        'sleeve_length': row.sleeve_length or '',
        'neck_women': row.neck_women or '',
        'sleeve_styling': row.sleeve_styling or '',
    }


def build_po_export_row_values(row: PORequestRow) -> list[Any]:
    field_map = _po_export_field_map(row)
    return [
        field_map['sku_id'],
        field_map['brand_name'],
        field_map['category_type'],
        field_map['styli_sku_id'],
        field_map['color'],
        field_map['size'],
        field_map['l1'],
        field_map['fibre_composition'],
        field_map['coo'],
        field_map['po_price'],
        field_map['osp_in_sar'],
        field_map['po_qty'],
        field_map['knitted_woven'],
        field_map['product_name'],
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        field_map['dress_print'],
        field_map['dress_length'],
        field_map['dress_shape'],
        field_map['sleeve_length'],
        field_map['neck_women'],
        field_map['sleeve_styling'],
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
    ]


def build_po_export_row_values_from_template(row: PORequestRow, columns: list[dict[str, Any]]) -> list[Any]:
    field_map = _po_export_field_map(row)
    return [field_map.get(str(column.get('source_field') or ''), '') for column in columns]


def _apply_header_styles(sheet) -> None:
    for column_index, (group_label, header_label) in enumerate(zip(PO_EXPORT_GROUP_ROW, PO_EXPORT_HEADERS), start=1):
        fill = PatternFill('solid', fgColor=HEADER_FILL_COLORS.get(group_label, 'E5E7EB'))
        group_cell = sheet.cell(row=1, column=column_index)
        header_cell = sheet.cell(row=2, column=column_index)

        group_cell.fill = fill
        group_cell.font = Font(bold=True, size=11)
        group_cell.border = HEADER_BORDER
        group_cell.alignment = HEADER_ALIGNMENT

        header_cell.fill = fill
        header_cell.font = Font(bold=True, size=10)
        header_cell.border = HEADER_BORDER
        header_cell.alignment = HEADER_ALIGNMENT

    start_index = 1
    current_label = PO_EXPORT_GROUP_ROW[0]
    for column_index, group_label in enumerate(PO_EXPORT_GROUP_ROW[1:], start=2):
        if group_label == current_label:
            continue
        if start_index != column_index - 1:
            sheet.merge_cells(start_row=1, start_column=start_index, end_row=1, end_column=column_index - 1)
        start_index = column_index
        current_label = group_label

    if start_index != len(PO_EXPORT_GROUP_ROW):
        sheet.merge_cells(start_row=1, start_column=start_index, end_row=1, end_column=len(PO_EXPORT_GROUP_ROW))


def _apply_column_widths(sheet) -> None:
    explicit_widths = {
        'A': 22,
        'B': 18,
        'C': 16,
        'D': 16,
        'E': 14,
        'F': 10,
        'G': 10,
        'H': 20,
        'I': 10,
        'J': 12,
        'K': 12,
        'L': 10,
        'M': 16,
        'N': 34,
    }

    for column_index, header_label in enumerate(PO_EXPORT_HEADERS, start=1):
        column_letter = get_column_letter(column_index)
        width = explicit_widths.get(column_letter, min(max(len(header_label) * 1.15, 13), 26))
        sheet.column_dimensions[column_letter].width = width


def build_po_export_csv(po_request: PORequest, template_config: dict[str, Any] | None = None) -> str:
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)
    ordered_rows = sorted(po_request.rows, key=lambda row: (row.row_index, row.created_at))

    if template_config is None:
        writer.writerow(PO_EXPORT_GROUP_ROW)
        writer.writerow(PO_EXPORT_HEADERS)
        for row in ordered_rows:
            writer.writerow(build_po_export_row_values(row))
        return output.getvalue()

    columns = list(template_config.get('columns') or [])
    layout = dict(template_config.get('layout') or {})
    prefix_rows = layout.get('prefix_rows')
    if isinstance(prefix_rows, list) and prefix_rows:
        for prefix_row in prefix_rows:
            writer.writerow(list(prefix_row) if isinstance(prefix_row, list) else [])
    else:
        writer.writerow([str(column.get('header') or '') for column in columns])

    for row in ordered_rows:
        writer.writerow(build_po_export_row_values_from_template(row, columns))
    return output.getvalue()


def build_po_export_xlsx(po_request: PORequest, template_config: dict[str, Any] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    ordered_rows = sorted(po_request.rows, key=lambda row: (row.row_index, row.created_at))

    if template_config is None:
        sheet.title = 'Women_SST_PO'
        sheet.append(PO_EXPORT_GROUP_ROW)
        sheet.append(PO_EXPORT_HEADERS)
        for row in ordered_rows:
            sheet.append(build_po_export_row_values(row))

        _apply_header_styles(sheet)
        _apply_column_widths(sheet)

        sheet.freeze_panes = 'A3'
        sheet.auto_filter.ref = f'A2:{get_column_letter(len(PO_EXPORT_HEADERS))}{sheet.max_row}'
        sheet.sheet_view.zoomScale = 85
        sheet.row_dimensions[1].height = 24
        sheet.row_dimensions[2].height = 42

        data_start_row = 3
        total_columns = len(PO_EXPORT_HEADERS)
    else:
        columns = list(template_config.get('columns') or [])
        layout = dict(template_config.get('layout') or {})
        prefix_rows = layout.get('prefix_rows')
        sheet.title = str(template_config.get('sheet_name') or 'Workbook')
        if isinstance(prefix_rows, list) and prefix_rows:
            for prefix_row in prefix_rows:
                sheet.append(list(prefix_row) if isinstance(prefix_row, list) else [])
        else:
            sheet.append([str(column.get('header') or '') for column in columns])

        for row in ordered_rows:
            sheet.append(build_po_export_row_values_from_template(row, columns))

        column_widths = layout.get('column_widths')
        if isinstance(column_widths, dict):
            for column_letter, width in column_widths.items():
                if isinstance(column_letter, str) and isinstance(width, (int, float)):
                    sheet.column_dimensions[column_letter].width = float(width)

        merged_ranges = layout.get('merged_ranges')
        if isinstance(merged_ranges, list):
            for merged_range in merged_ranges:
                if not isinstance(merged_range, str):
                    continue
                try:
                    sheet.merge_cells(merged_range)
                except ValueError:
                    continue

        header_row_index = int(template_config.get('header_row_index') or 1)
        freeze_panes = layout.get('freeze_panes')
        if isinstance(freeze_panes, str) and freeze_panes.strip():
            sheet.freeze_panes = freeze_panes
        else:
            sheet.freeze_panes = f'A{header_row_index + 1}'
        if columns:
            sheet.auto_filter.ref = f'A{header_row_index}:{get_column_letter(len(columns))}{sheet.max_row}'
        data_start_row = header_row_index + 1
        total_columns = len(columns)

    for row_index in range(data_start_row, sheet.max_row + 1):
        for column_index in range(1, total_columns + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = HEADER_BORDER
            cell.alignment = LEFT_BODY_ALIGNMENT if column_index in {1, 2, 8, 14} else BODY_ALIGNMENT

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
