import json
import re
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import urlopen
from uuid import uuid4
from xml.etree import ElementTree as ET

from sqlalchemy.orm import Session

from app.db.base import utcnow
from app.db.session import SessionLocal
from app.models.received_po import ReceivedPO, ReceivedPOLineItem

SIZE_ORDER = {'XS': 0, 'S': 1, 'M': 2, 'L': 3, 'XL': 4, 'XXL': 5, 'XXXL': 6}
LINE_ITEM_FIELD_ALIASES = {
    'brand_style_code': {'brand style code', 'brand_style_code', 'style code', 'style_code'},
    'styli_style_id': {'styli style id', 'styli_style_id', 'styli id', 'style id'},
    'model_number': {'model number', 'model_number', 'model no', 'model no.'},
    'option_id': {'option id', 'option_id', 'option'},
    'sku_id': {'sku id', 'sku_id', 'sku'},
    'color': {'color', 'colour'},
    'size': {'size'},
    'quantity': {'quantity', 'qty', 'pieces'},
    'po_price': {'po price', 'po_price', 'price', 'unit price', 'po rate'},
}


def _json_dumps(payload: dict[str, object]) -> str:
    return json.dumps(payload, separators=(',', ':'))


def _read_received_po_bytes(file_url: str) -> bytes:
    parsed = urlparse(file_url)
    if parsed.scheme in {'http', 'https'}:
        with urlopen(file_url, timeout=15) as response:
            return response.read()

    path_part = parsed.path if parsed.path else file_url
    if path_part.startswith('/static/'):
        local_path = Path(path_part.lstrip('/'))
    elif path_part.startswith('static/'):
        local_path = Path(path_part)
    else:
        return b''

    if not local_path.exists():
        api_root_local_path = Path(__file__).resolve().parents[2] / local_path
        if api_root_local_path.exists():
            return api_root_local_path.read_bytes()
        return b''
    return local_path.read_bytes()


def _extract_pdf_text(content: bytes) -> str:
    if not content:
        return ''

    try:
        import pdfplumber  # type: ignore
    except ImportError:
        pdfplumber = None

    if pdfplumber is not None:
        with pdfplumber.open(BytesIO(content)) as pdf:
            return '\n'.join((page.extract_text() or '').strip() for page in pdf.pages if page.extract_text())

    return content.decode('latin-1', errors='ignore')


def _extract_excel_rows(content: bytes) -> list[list[str]]:
    if not content:
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        load_workbook = None

    if load_workbook is not None:
        workbook = load_workbook(BytesIO(content), data_only=True)
        rows: list[list[str]] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                normalized_row = [str(cell).strip() if cell is not None else '' for cell in row]
                if any(normalized_row):
                    rows.append(normalized_row)
        return rows

    return _extract_xlsx_rows_fallback(content)


def _extract_xlsx_rows_fallback(content: bytes) -> list[list[str]]:
    namespace = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    rows: list[list[str]] = []

    with zipfile.ZipFile(BytesIO(content)) as archive:
        shared_strings: list[str] = []
        if 'xl/sharedStrings.xml' in archive.namelist():
            shared_root = ET.fromstring(archive.read('xl/sharedStrings.xml'))
            for item in shared_root.findall('main:si', namespace):
                text_fragments = [node.text or '' for node in item.findall('.//main:t', namespace)]
                shared_strings.append(''.join(text_fragments).strip())

        worksheet_names = sorted(
            name for name in archive.namelist() if name.startswith('xl/worksheets/sheet') and name.endswith('.xml')
        )
        for worksheet_name in worksheet_names:
            worksheet_root = ET.fromstring(archive.read(worksheet_name))
            for row_node in worksheet_root.findall('.//main:sheetData/main:row', namespace):
                current_row: list[str] = []
                for cell in row_node.findall('main:c', namespace):
                    cell_type = cell.attrib.get('t')
                    value_node = cell.find('main:v', namespace)
                    inline_text_node = cell.find('main:is/main:t', namespace)
                    value = ''
                    if inline_text_node is not None:
                        value = inline_text_node.text or ''
                    elif value_node is not None and value_node.text is not None:
                        raw_value = value_node.text
                        if cell_type == 's':
                            try:
                                value = shared_strings[int(raw_value)]
                            except (ValueError, IndexError):
                                value = raw_value
                        else:
                            value = raw_value
                    current_row.append(value.strip())
                if any(current_row):
                    rows.append(current_row)
    return rows


def _normalize_header(value: str) -> str:
    return re.sub(r'\s+', ' ', value.strip().lower())


def _find_header_mapping(rows: list[list[str]]) -> tuple[int, dict[str, int]] | tuple[None, dict[str, int]]:
    for row_index, row in enumerate(rows):
        header_map: dict[str, int] = {}
        for column_index, raw_cell in enumerate(row):
            normalized = _normalize_header(raw_cell)
            for field_name, aliases in LINE_ITEM_FIELD_ALIASES.items():
                if normalized in aliases and field_name not in header_map:
                    header_map[field_name] = column_index
        if {'brand_style_code', 'sku_id', 'quantity'} <= set(header_map):
            return row_index, header_map
    return None, {}


def _parse_decimal(value: str) -> float | None:
    cleaned = re.sub(r'[^0-9.]', '', value)
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_int(value: str) -> int:
    cleaned = re.sub(r'[^0-9-]', '', value)
    if not cleaned:
        return 0
    try:
        return max(0, int(cleaned))
    except ValueError:
        return 0


def _extract_line_items_from_rows(rows: list[list[str]]) -> list[dict[str, object]]:
    header_row_index, header_map = _find_header_mapping(rows)
    if header_row_index is None:
        return []

    items: list[dict[str, object]] = []
    for row in rows[header_row_index + 1 :]:
        if len(row) <= max(header_map.values()):
            continue

        def _value(field_name: str) -> str:
            column_index = header_map.get(field_name)
            if column_index is None or column_index >= len(row):
                return ''
            return row[column_index].strip()

        brand_style_code = _value('brand_style_code')
        sku_id = _value('sku_id')
        quantity = _parse_int(_value('quantity'))
        if not brand_style_code or not sku_id:
            continue

        items.append(
            {
                'brand_style_code': brand_style_code,
                'styli_style_id': _value('styli_style_id') or None,
                'model_number': _value('model_number') or None,
                'option_id': _value('option_id') or None,
                'sku_id': sku_id,
                'color': _value('color') or None,
                'size': _value('size') or None,
                'quantity': quantity,
                'po_price': _parse_decimal(_value('po_price')),
            }
        )
    return items


def _extract_po_number(raw_text: str) -> str | None:
    patterns = [
        r'\b(STY-\d{4}-\d{3,})\b',
        r'po\s*(?:number|no\.?)\s*[:\-]?\s*([A-Za-z0-9\-\/]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, raw_text, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return None


def _extract_po_date(raw_text: str) -> datetime | None:
    patterns = [
        r'po\s*date\s*[:\-]?\s*(\d{4}-\d{2}-\d{2})',
        r'po\s*date\s*[:\-]?\s*(\d{2}/\d{2}/\d{4})',
        r'po\s*date\s*[:\-]?\s*(\d{2}-\d{2}-\d{4})',
    ]
    for pattern in patterns:
        match = re.search(pattern, raw_text, flags=re.IGNORECASE)
        if not match:
            continue
        value = match.group(1).strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(value, fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
    return None


def _guess_distributor(raw_text: str) -> str:
    return 'Styli' if 'styli' in raw_text.lower() else 'Unknown'


def _sort_line_items(items: list[dict[str, object]]) -> list[dict[str, object]]:
    return sorted(
        items,
        key=lambda item: (
            str(item.get('brand_style_code') or ''),
            str(item.get('option_id') or ''),
            SIZE_ORDER.get(str(item.get('size') or '').upper(), 999),
            str(item.get('sku_id') or ''),
        ),
    )


def parse_received_po_file(file_url: str) -> dict[str, object]:
    file_bytes = _read_received_po_bytes(file_url)
    suffix = Path(urlparse(file_url).path or file_url).suffix.lower()

    rows: list[list[str]] = []
    raw_text = ''
    if suffix in {'.xlsx', '.xls'}:
        rows = _extract_excel_rows(file_bytes)
        raw_text = '\n'.join(' | '.join(row) for row in rows)
    else:
        raw_text = _extract_pdf_text(file_bytes)

    items = _extract_line_items_from_rows(rows)
    parsed_payload = {
        'po_number': _extract_po_number(raw_text),
        'po_date': _extract_po_date(raw_text).isoformat() if _extract_po_date(raw_text) else None,
        'distributor': _guess_distributor(raw_text),
        'line_items': _sort_line_items(items),
        'raw_text_preview': raw_text[:1500],
    }
    return parsed_payload


def process_received_po_parse_job(received_po_id: str) -> None:
    db: Session = SessionLocal()
    record: ReceivedPO | None = None
    try:
        record = db.query(ReceivedPO).filter(ReceivedPO.id == received_po_id).first()
        if record is None:
            return

        record.status = 'parsing'
        db.commit()

        parsed_payload = parse_received_po_file(record.file_url)
        line_items = parsed_payload.get('line_items')
        if not isinstance(line_items, list) or len(line_items) == 0:
            record.status = 'failed'
            record.raw_extracted_json = _json_dumps(
                {
                    **parsed_payload,
                    'parse_error': 'No line items could be extracted from the received PO.',
                }
            )
            db.commit()
            return

        db.query(ReceivedPOLineItem).filter(ReceivedPOLineItem.received_po_id == record.id).delete()

        for item in line_items:
            if not isinstance(item, dict):
                continue
            db.add(
                ReceivedPOLineItem(
                    id=str(uuid4()),
                    received_po_id=record.id,
                    brand_style_code=str(item.get('brand_style_code') or '').strip(),
                    styli_style_id=str(item.get('styli_style_id') or '').strip() or None,
                    model_number=str(item.get('model_number') or '').strip() or None,
                    option_id=str(item.get('option_id') or '').strip() or None,
                    sku_id=str(item.get('sku_id') or '').strip(),
                    color=str(item.get('color') or '').strip() or None,
                    size=str(item.get('size') or '').strip() or None,
                    quantity=max(0, int(item.get('quantity') or 0)),
                    po_price=float(item['po_price']) if item.get('po_price') is not None else None,
                )
            )

        po_date_value = parsed_payload.get('po_date')
        parsed_po_date = None
        if isinstance(po_date_value, str) and po_date_value.strip():
            try:
                parsed_po_date = datetime.fromisoformat(po_date_value.replace('Z', '+00:00'))
            except ValueError:
                parsed_po_date = None

        record.po_number = str(parsed_payload.get('po_number') or '').strip() or record.po_number
        record.po_date = parsed_po_date or record.po_date
        record.distributor = str(parsed_payload.get('distributor') or '').strip() or record.distributor
        record.raw_extracted_json = _json_dumps(parsed_payload)
        record.updated_at = utcnow()
        record.status = 'parsed'
        db.commit()
    except Exception as exc:
        if record is not None:
            record.status = 'failed'
            record.raw_extracted_json = _json_dumps({'parse_error': str(exc)[:500]})
            db.commit()
    finally:
        db.close()
